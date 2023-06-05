from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
import json
from django.http import JsonResponse
from authentication.models import Ticket




class ExcelTemplateGenerator:
    def generate_template(self, headers):
        wb = openpyxl.Workbook()
        sheet = wb.active

        for i, header in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            sheet.column_dimensions[col_letter].width = 15
            sheet.cell(row=1, column=i, value=header)

        return wb

    def download_template(self, wb, filename):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


def download_userprofiles_template(request):
    generator = ExcelTemplateGenerator()
    headers = ["Username", "First Name", "Last Name", "Email", "Password", "User Category", "User Country", "Track Code", "User University", "identifier"]
    wb = generator.generate_template(headers)
    return generator.download_template(wb, 'userprofiles_template.xlsx')

def download_event_template(request):
    generator = ExcelTemplateGenerator()
    headers = ['eventCode', 'eventTheme', 'eventStartTime', 'eventEndTime', 'conference_id', 'keynoteSpeaker', 'eventRoom']
    wb = generator.generate_template(headers)
    return generator.download_template(wb, 'events_template.xlsx')


def download_paper_template(request):
    generator = ExcelTemplateGenerator()
    headers = ['user', 'paperID', 'paperTitle']
    wb = generator.generate_template(headers)
    return generator.download_template(wb, 'papers_template.xlsx')
    
    


def process_ticket(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        ticket_id = data.get('ticket_id')
        print("Ticket ID:", ticket_id)
        try:
            ticket = Ticket.objects.get(ticket_id=ticket_id)
            ticket.checkin = True
            ticket.save()
            print("Check-in updated successfully")
            return JsonResponse({"success": True})
        except Ticket.DoesNotExist:
            print("Ticket not found")
            return JsonResponse({"success": False})
    return JsonResponse({"success": False})