from django.shortcuts import redirect, render
from django.db.models import Count
from django.contrib.auth.models import User
from django.contrib import messages
import openpyxl
from authentication.models import Conference, UserProfile, Event, Paper, Track, Ticket
from .forms import AdminEventFileUploadForm, AdminUserProfileFileUploadForm, AdminPaperFileUploadForm
from django.http import HttpResponseForbidden, JsonResponse
from django.views.decorators.csrf import csrf_exempt


def home_dashboard(request):
    conferences = Conference.objects.all()

    # Prepare data for chart
    chart_data = []
    for conference in conferences:
        user_profiles = UserProfile.objects.filter(
            tickets__track__conference=conference)
        checked_tickets = Ticket.objects.filter(
            track__conference=conference, checkin=True)
        unchecked_tickets = Ticket.objects.filter(
            track__conference=conference, checkin=False)

        checkin_count = checked_tickets.count()
        unchecked_count = unchecked_tickets.count()
        total_participants = checkin_count + unchecked_count
        checkin_percentage = round(
            (checkin_count / total_participants) * 100, 1) if total_participants > 0 else 0

        context = {
            'checked_tickets': checked_tickets,
            'unchecked_tickets': unchecked_tickets,
            'checkin_count': checkin_count,
            'unchecked_count': unchecked_count,
            'checkin_percentage': checkin_percentage
        }
        country_count = user_profiles.values(
            'userCountry').annotate(count=Count('userCountry'))
        chart_data.append({
            'conference': conference.conferenceName,
            'country_count': list(country_count),
        })

    return render(request, 'custom_admin_dashboard/home_dashboard.html',  {
        'chart_data': chart_data, 'context': context,
    })

# custom_admin_dashboard/admin.py


def scanner(request):
    if request.user.is_staff:
        return render(request, 'scanner.html')
    return HttpResponseRedirect('/admin/')


# @csrf_exempt
# def process_qr_code(request):
#     if not request.user.is_staff:
#         return HttpResponseForbidden()

#     if request.method == 'POST':
#         ticket_id = request.POST.get('ticket_id')
#         try:
#             ticket = Ticket.objects.get(id=ticket_id)
#             ticket.checkin = True
#             ticket.save()
#             return JsonResponse({'status': 'success'})
#         except Ticket.DoesNotExist:
#             return JsonResponse({'status': 'error', 'message': 'Ticket not found'})
#     else:
#         return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


def upload_event_file(request):
    if request.method == 'POST':
        form = AdminEventFileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            event_file = request.FILES['event_file']
            wb = openpyxl.load_workbook(event_file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2):  # Skip header row
                event = Event(
                    eventCode=row[0].value,
                    eventTheme=row[1].value,
                    eventStartTime=row[2].value,
                    eventEndTime=row[3].value,
                    # Assuming conference_id is stored in the Excel file
                    conference_id=row[4].value,
                    keynoteSpeaker=row[5].value,
                    eventRoom=row[6].value,
                )
                event.save()
            # Redirect to the Event list page in the admin site
            return redirect('admin:authentication_event_changelist')
    else:
        form = AdminEventFileUploadForm()
    return render(request, 'admin/upload_event_file.html', {'form': form})


def upload_userprofiles_file(request):
    if request.method == 'POST':
        form = AdminUserProfileFileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            user_profiles_file = request.FILES['user_profiles_file']
            # Assuming conferenceCode is stored in the Excel file
            wb = openpyxl.load_workbook(user_profiles_file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2):  # Skip header row
                username = row[0].value
                fname = row[1].value
                lname = row[2].value
                email = row[3].value
                password = row[4].value
                user_category = row[5].value
                user_country = row[6].value
                trackCode = row[7].value
                user_univeristy = row[8].value
                identifier = row[9].value
                if username == None:
                    break
                user, created = User.objects.get_or_create(
                    username=username,
                    email=email,
                )
                user.set_password(password)
                user.first_name = fname
                user.last_name = lname
                userprofile = UserProfile.objects.get(user=user)
                userprofile.userCategory = user_category
                userprofile.userCountry = user_country
                userprofile.userUniversity = user_univeristy
                userprofile.identifier = identifier
                # only update if the track exists
                if not Track.objects.filter(trackCode=trackCode).exists():
                    messages.error(
                        request, f"Track {trackCode} does not exist.")
                else:
                    if Ticket.objects.filter(user=user, track__trackCode=trackCode).exists():
                        messages.error(
                            request, f"User {user.username} already has a ticket for track {trackCode}.")
                    else:
                        ticket = Ticket.objects.create(
                            user=user,
                            track=Track.objects.get(trackCode=trackCode),
                        )
                        ticket.save()  # it also generate a ticket id
                        user.userprofile.tickets.add(ticket)
                user.save()  # Save the both user and UserProfile model instance

            messages.success(
                request, "User profiles have been successfully imported.")

            # URL pattern of admin follows the singular form of the model name
            return redirect('admin:authentication_userprofile_changelist')
    else:
        form = AdminUserProfileFileUploadForm()

    return render(request, 'admin/upload_userprofiles_file.html', {'form': form})


def upload_paper_file(request):
    if request.method == 'POST':
        form = AdminPaperFileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            paper_file = request.FILES['paper_file']
            wb = openpyxl.load_workbook(paper_file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2):  # Skip header row
                user_email = row[0].value
                user = User.objects.get(email=user_email)
                paper = Paper(
                    user=user,
                    paperID=row[1].value,
                    paperTitle=row[2].value,
                )
                paper.save()
            return redirect('admin:authentication_paper_changelist')
    else:
        form = AdminPaperFileUploadForm()
    return render(request, 'admin/upload_paper_file.html', {'form': form})
